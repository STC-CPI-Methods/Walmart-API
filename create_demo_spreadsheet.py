import json
import pandas as pd
import openpyxl as op

"""
This script parses response files generated by the Walmart API and outputs the parsed responses to separate sheets in an Excel workbook.
Note that this script does not generate an Excel file. It must be created beforehanded, along with the sheets.
This script was written for demo purposes only and was not designed to parse generic responses.
"""


class APIResponse:
    """
    This class stores the attributes of a response and instructs the script how the response should be parsed.
    
    Instance Variables:
        fileName: name of the input JSON file
        sheetName: name of the Excel worksheet to which the parsed response will be outputted
        startRow: first row of the worksheet at which the response can be written
        canDump: True if the JSON response can simply be dumped into a worksheet. False if the JSON response must be parsed node by node.
        dumpKey: Identifies the part of the JSON response that will be dumped.
        attributes: list of attributes that will be printed at the top of the sheet.
    """
    
    def __init__(self, argFileName, argSheetName, argStartRow, argCanDump, argDumpKey, argAttributes):
        self.fileName = argFileName
        self.sheetName = argSheetName
        self.startRow = argStartRow
        self.canDump = argCanDump
        self.dumpKey = argDumpKey
        self.attributes = argAttributes



def printTree(obj, sheet, rowIndex, colIndex):
    """
    This method parses a JSON response and prints the contents of each node to a worksheet.
    
    Arguments:
        obj: the object to be parsed; can either be a dictionary, a list or a single value.
        sheet: an Excel worksheet object; the parsed response is outputted to this sheet.
        rowIndex: index of the current row of the response.
        colIndex: index of the current column of the response.
    """
    
    rowIndex = rowIndex + 1
    
    # Case 1: obj is a dictionary
    if type(obj) is dict:
        
        # Iterate over all key-value pairs
        for key, value in obj.items():
            sheet.cell(row=rowIndex, column=colIndex).value = key # Write key to cell
            rowIndex = printTree(value, sheet, rowIndex, colIndex + 1) # recursive call to printTree
        return rowIndex
    
    # Case 2: obj is a list
    elif type(obj) is list:
        
        # Iterate over all list items.
        for i in range(0, len(obj)):
            rowIndex = printTree(obj[i], sheet, rowIndex, colIndex) # recursive call to printTree
        return rowIndex
    
    # Case 3: obj is a single value
    else:
        rowIndex = rowIndex - 1
        if (rowIndex == 0):
            rowIndex = 1
        sheet.cell(row=rowIndex, column=colIndex).value = obj # Write value to cell
        rowIndex = rowIndex + 1
        return rowIndex
        
            
# Construct the API Response objects
search = APIResponse('search.txt', 'Search', 7, True, 'items',
                                ['query', 'sort', 'responseGroup', 'totalResults',
                                 'start', 'numItems'])
productLookup = APIResponse('productLookup.txt', 'Product Lookup', 1, False, '', [])
paginatedProducts = APIResponse('paginatedProducts.txt', 'Paginated Products', 5, True, 'items',
                                ['category', 'format', 'nextPage', 'totalPages'])
taxonomy = APIResponse('taxonomy.txt', 'Taxonomy', 1, False, '', [])

# List of API Response objects
apiResponses = [search, productLookup, paginatedProducts, taxonomy]


# Initialize the workbook.
excelFileName = 'WalmartAPIDemo.xlsx'
writer = pd.ExcelWriter(excelFileName, engine='openpyxl')
book = op.load_workbook(excelFileName, read_only=False)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


# Overwrite the worksheets
for i in range(0, 4):
    
    # Load the JSON response
    with open(apiResponses[i].fileName, encoding='utf-8') as file:
        fileString = file.read()
    dataDict = json.loads(fileString)
    
    # Print the desired attributes at the top of the worksheet
    if (len(apiResponses[i].attributes) != 0):
        sheet = book[apiResponses[i].sheetName]
        for j in range(1, len(apiResponses[i].attributes) + 1):
            sheet.cell(row=j, column=1).value = apiResponses[i].attributes[j - 1]
            sheet.cell(row=j, column=2).value = dataDict[apiResponses[i].attributes[j - 1]]
    
    # Print the response (or the body of the response) into the worksheet
    
    # Case 1: Response is dumped.
    if (apiResponses[i].canDump == True):
        
        # Case 1.1: No dump key; entire response is dumped.
        if (apiResponses[i].dumpKey != ''):
            df = pd.DataFrame(data=dataDict[apiResponses[i].dumpKey])
            
        # Case 1.2: Has a dump key: only the portion of the response corresponding to the key is dumped.
        else:
            df = pd.DataFrame(data=dataDict)
        
        df.to_excel(writer, sheet_name=apiResponses[i].sheetName, startrow=apiResponses[i].startRow)
        
    # Case 2: Response is parsed.
    else:
        sheet = book[apiResponses[i].sheetName]
        printTree(dataDict, sheet, apiResponses[i].startRow - 1, 1)
    
    
# Save the workbook.
writer.save()